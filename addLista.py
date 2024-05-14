import customtkinter as ctk
from tkinter import messagebox
from datetime import datetime
import openpyxl
from openpyxl import Workbook, load_workbook

def limpar_campos():
    # Limpa todos os campos de entrada
    for entry in [nome_entry, telefone_entry, telefone1_entry, telefone_residencial_entry, idade_entry, data_nascimento_entry, email_entry]:
        entry.delete(0, 'end')
    
    # Restaura os números predefinidos nos campos de telefone
    telefone_entry.insert(0, '55(11)9')
    telefone1_entry.insert(0, '55(11)9')
    telefone_residencial_entry.insert(0, '55(11)9')

def adicionar_contato():
    # Obter os valores dos campos
    nome = nome_entry.get() or 'Null'
    telefone = telefone_entry.get()
    telefone1 = telefone1_entry.get()
    telefone_residencial = telefone_residencial_entry.get()
    idade = idade_entry.get()
    data_nascimento = data_nascimento_entry.get()
    email = email_entry.get() or 'Null'

    # Formatar os telefones
    telefone = formatar_telefone(telefone)
    telefone1 = formatar_telefone(telefone1)
    telefone_residencial = formatar_telefone(telefone_residencial)

    # Verificar se a idade não foi informada, mas a data de nascimento foi
    if idade == '' and data_nascimento != '':
        idade = calcular_idade(data_nascimento)

    # Abrir o arquivo Excel (se já existir) ou criar um novo
    try:
        wb = load_workbook('contatos.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(['Nome', 'Telefone', 'Telefone1', 'Telefone Residencial', 'Idade', 'Data de Nascimento', 'Email'])

    # Adicionar o novo contato na próxima linha disponível
    next_row = sheet.max_row + 1
    sheet.append([nome, telefone, telefone1, telefone_residencial, idade, data_nascimento, email])
    wb.save('contatos.xlsx')

    # Mostrar mensagem de sucesso e limpar campos
    messagebox.showinfo("Sucesso", "Contato adicionado com sucesso!")
    limpar_campos()

def calcular_idade(data_nascimento):
    # Calcular a idade com base na data de nascimento
    data_atual = datetime.now()
    ano_atual = data_atual.year
    try:
        ano_nascimento = int(data_nascimento[-4:])
        idade = ano_atual - ano_nascimento
        return idade
    except ValueError:
        return 'Null'

def formatar_telefone(telefone):
    # Se o telefone estiver vazio ou 'Null', retorna 'Null'
    if not telefone or telefone == 'Null':
        return 'Null'
    
    # Remove caracteres não numéricos
    telefone = ''.join(c for c in telefone if c.isdigit())

    # Formatação do telefone se tiver 9 dígitos
    if len(telefone) == 9:
        telefone = f"55(11)9{telefone[:4]}-{telefone[4:]}"
    return telefone

def add_dash(event):
    # Adicionar traço ao telefone ao alcançar 9 dígitos
    current_text = event.widget.get()
    digit_count = sum(c.isdigit() for c in current_text)
    if digit_count == 9 and '-' not in current_text:
        event.widget.insert('end', '-')

root = ctk.CTk()
root.title("Adicionar Contato")
root.geometry("800x600")

ctk_label = ctk.CTkLabel
ctk_entry = ctk.CTkEntry
ctk_button = ctk.CTkButton

# Função para adicionar traço ao telefone ao alcançar 9 dígitos
def bind_add_dash(entry):
    entry.bind('<KeyRelease>', add_dash)

# Criar os elementos da interface
ctk_label(root, text="Nome:").grid(row=0, column=0, padx=5, pady=5)
nome_entry = ctk_entry(root, width=250)
nome_entry.grid(row=0, column=1, padx=5, pady=5)

ctk_label(root, text="Telefone:").grid(row=1, column=0, padx=5, pady=5)
telefone_entry = ctk_entry(root, width=250)
telefone_entry.insert(0, '55(11)9')
telefone_entry.grid(row=1, column=1, padx=5, pady=5)
bind_add_dash(telefone_entry)

ctk_label(root, text="Telefone1:").grid(row=2, column=0, padx=5, pady=5)
telefone1_entry = ctk_entry(root, width=250)
telefone1_entry.grid(row=2, column=1, padx=5, pady=5)
bind_add_dash(telefone1_entry)

ctk_label(root, text="Telefone Residencial:").grid(row=3, column=0, padx=5, pady=5)
telefone_residencial_entry = ctk_entry(root, width=250)
telefone_residencial_entry.grid(row=3, column=1, padx=5, pady=5)
bind_add_dash(telefone_residencial_entry)

ctk_label(root, text="Idade:").grid(row=4, column=0, padx=5, pady=5)
idade_entry = ctk_entry(root, width=50)
idade_entry.grid(row=4, column=1, padx=5, pady=5)

ctk_label(root, text="Data de Nascimento:").grid(row=5, column=0, padx=5, pady=5)
data_nascimento_entry = ctk_entry(root, width=250)
data_nascimento_entry.grid(row=5, column=1, padx=5, pady=5)
data_nascimento_entry.bind('<KeyRelease>', lambda event: data_nascimento_entry.insert('end', '/') if len(data_nascimento_entry.get()) in [2, 5] else None)


ctk_label(root, text="Email:").grid(row=6, column=0, padx=5, pady=5)
email_entry = ctk_entry(root, width=250)
email_entry.grid(row=6, column=1, padx=5, pady=5)

# Botão para adicionar contato
adicionar_button = ctk_button(root, text="Adicionar Contato", command=adicionar_contato)
adicionar_button.grid(row=7, columnspan=2, pady=10)

root.mainloop()
